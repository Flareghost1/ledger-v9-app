# -*- coding: utf-8 -*-
"""
기존 자산 명세서(xlsx) → v7 거래원장(transactions.csv) + 마스터 SEED 마이그레이션.
SPEC §7 Step1(기초잔고 SEED) + Step2(필수 소급: 증여·NVDA·브라질·에너지) 반영.

실행:  python migrate_seed.py "경로/명세서.xlsx"
       (경로 생략 시 ../나의 자산 또는 ../자산현황 의 최신 xlsx 자동 탐색)
결과:  data/transactions.csv, data/assets.csv, data/accounts.csv
※ transactions.csv는 이후 손으로 편집하는 유일한 원장이 된다(다시 덮어쓰지 말 것).
"""
import sys, csv, re
from pathlib import Path
import openpyxl
sys.path.insert(0, str(Path(__file__).resolve().parent / "src"))
from asset_meta import match_meta, CASH

BASE = Path(__file__).resolve().parent
DATA = BASE / "data"
FX_USD = 1520.4  # 명세서 기준일 환율

# 알려진 취득일 (모르면 기준일 2026-06-16). SPEC Step2 필수 소급.
ACQ_DATE = {
    "NVDA": "2026-06-18", "LNG": "2026-06-18", "XOM": "2026-06-18", "COP": "2026-06-18",
    "BOND_FX.브라질2037": "2026-06-18", "BOND_FX.브라질2035신규": "2026-06-18",
    "BOND_FX.브라질2033": "2026-06-18", "BOND_FX.브라질2031": "2026-06-18",
    "BOND_FX.브라질2029": "2026-06-18",
    "088980.KS": "2026-06-15", "483280.KS": "2026-06-15", "494300.KS": "2026-06-15",
    "ELS.교보ELB12532": "2026-06-12",
}
SEED_DATE = "2026-06-16"

def norm_acct(a):
    if a is None: return ""
    return re.sub(r"\s+", "", str(a))

def find_xlsx():
    if len(sys.argv) > 1:
        return Path(sys.argv[1])
    for folder in ("나의 자산", "자산현황"):
        d = BASE.parent / folder
        if d.is_dir():
            cands = sorted((p for p in d.glob("*통합_자산*.xlsx") if not p.name.startswith("~$")),
                           key=lambda p: p.name, reverse=True)
            if cands: return cands[0]
    raise SystemExit("명세서 xlsx를 찾을 수 없습니다. 경로를 인자로 주세요.")

def sheet_by(wb, key):
    for s in wb.sheetnames:
        if key in s: return wb[s]
    return None

def main():
    src = find_xlsx()
    print("소스:", src)
    wb = openpyxl.load_workbook(src, data_only=True)
    txns = []
    assets = {}   # asset_id -> dict
    seq = [0]
    def tid(date):
        seq[0] += 1
        return f"{date.replace('-','')}-{seq[0]:03d}"

    def add_asset(meta):
        aid, name, cls, psrc, ccy = meta
        assets.setdefault(aid, dict(asset_id=aid, name_kr=name, asset_class=cls,
                                    price_source=psrc, ccy=ccy, tax_exempt="N"))

    def add_seed_buy(owner, account, meta, qty, price, ccy, note, date=None, tag="SEED"):
        aid, name, cls, psrc, mccy = meta
        add_asset(meta)
        d = date or ACQ_DATE.get(aid, SEED_DATE)
        fx = FX_USD if ccy == "USD" else 1
        txns.append(dict(txn_id=tid(d), date=d, settle_date="", owner=owner, account=norm_acct(account),
                         asset_id=aid, asset_class=cls, type="BUY", qty=qty, price=price, ccy=ccy,
                         fx=fx, amount=round(qty*price, 4), fee=0, tax=0, cash_account="", link_id="",
                         tag=tag, note=note))

    def add_cash(owner, account, usd, krw, note):
        # USD 잔고가 있으면 CASH.USD, 아니면 CASH.KRW
        if usd:
            aid = "CASH.USD"
            assets.setdefault(aid, dict(asset_id=aid, name_kr="달러 예수금", asset_class=CASH,
                                        price_source="manual", ccy="USD", tax_exempt="N"))
            txns.append(dict(txn_id=tid(SEED_DATE), date=SEED_DATE, settle_date="", owner=owner,
                             account=norm_acct(account), asset_id=aid, asset_class=CASH, type="DEPOSIT",
                             qty="", price="", ccy="USD", fx=FX_USD, amount=round(usd,2), fee=0, tax=0,
                             cash_account=norm_acct(account), link_id="", tag="SEED", note=note))
        else:
            aid = "CASH.KRW"
            assets.setdefault(aid, dict(asset_id=aid, name_kr="원화 예수금", asset_class=CASH,
                                        price_source="manual", ccy="KRW", tax_exempt="N"))
            txns.append(dict(txn_id=tid(SEED_DATE), date=SEED_DATE, settle_date="", owner=owner,
                             account=norm_acct(account), asset_id=aid, asset_class=CASH, type="DEPOSIT",
                             qty="", price="", ccy="KRW", fx=1, amount=round(krw,0), fee=0, tax=0,
                             cash_account=norm_acct(account), link_id="", tag="SEED", note=note))

    # ── 시트2: 본인 금융자산 ──
    ws = sheet_by(wb, "금융자산")
    cur_sec = ""
    for row in ws.iter_rows(values_only=True):
        c0, c1 = row[0], row[1]
        if isinstance(c0, str) and c0.startswith("■"):
            cur_sec = c0
            continue
        if c1 is None or (isinstance(c1, str) and ("소계" in c1 or c1 in ("종목","항목","상품"))):
            continue
        if c0 in (None,) and c1 is None:
            continue
        acct, name = c0, str(c1)
        qty, p3, p4, val = row[2], row[3], row[4], row[5]
        if "현금성" in cur_sec:
            usd = qty if isinstance(qty, (int,float)) else None
            add_cash("본인", acct, usd, val if isinstance(val,(int,float)) else 0, name)
            continue
        meta = match_meta(name)
        if meta is None:
            continue
        aid, _, cls, psrc, mccy = meta
        if psrc.startswith("yfinance") and isinstance(qty,(int,float)) and isinstance(p3,(int,float)):
            add_seed_buy("본인", acct, meta, qty, p3, mccy, name)
        else:
            # 채권/펀드/ELS/연금 등 = 1좌 lump, price=평가액(원)
            if isinstance(val,(int,float)) and val > 0:
                add_seed_buy("본인", acct, meta, 1, round(val,0), "KRW", name)

    # ── 시트3: 증여·자녀자산 ──
    ws3 = sheet_by(wb, "증여")
    if ws3:
        cur_owner = None; carry = ""
        for row in ws3.iter_rows(values_only=True):
            c0 = row[0]
            if isinstance(c0, str) and c0.startswith("■"):
                if "배우자" in c0: cur_owner, carry = "배우자", ""
                elif "장남" in c0: cur_owner = "장남"; carry = "2027-05-18" if "이월과세" in c0 else ""
                elif "차남" in c0: cur_owner = "차남"; carry = "2027-05-18" if "이월과세" in c0 else ""
                continue
            if not cur_owner: continue
            owner, acct, name = row[0], row[1], row[2]
            qty, avg, curp, fxr, val = row[3], row[4], row[5], row[6], row[7]
            if name is None or (isinstance(name,str) and ("소계" in str(name) or name=="종목")):
                continue
            meta = match_meta(str(name))
            if meta is None: continue
            aid, _, cls, psrc, mccy = meta
            if "AAPL" in str(name) and "증여" in str(name) or (aid=="AAPL" and cur_owner in ("배우자","장남","차남")):
                # 증여 수증 (GIFT_IN) — 이월과세 시계 시작
                gdate = "2024-09-19" if cur_owner=="배우자" else "2026-05-18"
                add_asset(meta)
                fx = fxr if isinstance(fxr,(int,float)) else FX_USD
                note = f"{cur_owner} AAPL 증여 수증" + (f" · 이월과세 만료 {carry}" if carry else " · 이월과세 무관")
                txns.append(dict(txn_id=tid(gdate), date=gdate, settle_date="", owner=cur_owner,
                                 account=norm_acct(acct), asset_id="AAPL", asset_class="US_STOCK",
                                 type="GIFT_IN", qty=qty, price=avg, ccy="USD", fx=fx,
                                 amount=round((qty or 0)*(avg or 0),2), fee=0, tax=0, cash_account="",
                                 link_id="", tag="GIFT", note=note + (f" | carryover_expiry={carry}" if carry else "")))
            elif isinstance(qty,(int,float)) and isinstance(avg,(int,float)):
                ccy2 = "USD" if psrc.startswith("yfinance") and not aid.endswith((".KS",".KQ")) else "KRW"
                add_seed_buy(cur_owner, acct, meta, qty, avg, ccy2, f"{cur_owner} 운용")

    # ── 시트4: 실물·기타 ──
    ws4 = sheet_by(wb, "실물")
    if ws4:
        for row in ws4.iter_rows(values_only=True):
            cat, item, dref, val = row[0], row[1], row[2], row[3]
            if item is None or not isinstance(val,(int,float)) or val<=0: continue
            if isinstance(item,str) and "소계" in item: continue
            meta = match_meta(str(item))
            if meta is None: continue
            add_seed_buy("본인", "실물", meta, 1, round(val,0), "KRW", str(item))

    # ── 저장 ──
    DATA.mkdir(exist_ok=True)
    cols = ["txn_id","date","settle_date","owner","account","asset_id","asset_class","type","qty",
            "price","ccy","fx","amount","fee","tax","cash_account","link_id","tag","note"]
    with open(DATA/"transactions.csv","w",encoding="utf-8-sig",newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols); w.writeheader()
        for t in txns: w.writerow(t)
    acols = ["asset_id","name_kr","asset_class","price_source","ccy","tax_exempt","div_per_share","div_months"]
    # 브라질국채 비과세 표시
    for aid,a in assets.items():
        if a["asset_class"]=="BOND_FX": a["tax_exempt"]="Y"
        a.setdefault("div_per_share",""); a.setdefault("div_months","")
    with open(DATA/"assets.csv","w",encoding="utf-8-sig",newline="") as f:
        w = csv.DictWriter(f, fieldnames=acols); w.writeheader()
        for a in assets.values(): w.writerow({k:a.get(k,"") for k in acols})
    # 계좌 마스터
    accts = {}
    for t in txns:
        a = t["account"]
        if a and a not in accts:
            owner = t["owner"]
            broker = "삼성" if a.startswith("삼성") else ("신한" if a.startswith("신한") else
                     ("E*TRADE" if "TRADE" in a.upper() else ("교보" if "교보" in a else "기타")))
            atype = "RIA" if "RIA" in a else ("신탁" if "신탁" in a else ("연금" if broker=="교보" else "일반"))
            restr = "RIA_NO_WITHDRAW_1Y;KR_STOCK_ONLY" if "RIA" in a else ""
            accts[a] = dict(account=a, owner=owner, broker=broker, account_type=atype, restrictions=restr)
    with open(DATA/"accounts.csv","w",encoding="utf-8-sig",newline="") as f:
        w = csv.DictWriter(f, fieldnames=["account","owner","broker","account_type","restrictions"])
        w.writeheader()
        for a in accts.values(): w.writerow(a)

    print(f"거래 {len(txns)}건, 자산 {len(assets)}종, 계좌 {len(accts)}개 생성 → {DATA}")

if __name__ == "__main__":
    main()
